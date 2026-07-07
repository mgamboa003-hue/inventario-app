# app.py -- Inventario Wintec v3
import io
import os
import zipfile
from collections import defaultdict
from datetime import datetime, timedelta
from functools import wraps

import bcrypt
from dotenv import load_dotenv
from flask import (Flask, Response, flash, jsonify, redirect, render_template,
                   request, send_file, session, url_for)
from flask_wtf.csrf import CSRFProtect
from werkzeug.utils import secure_filename

from db import get_db_connection, init_db, p, USE_POSTGRES
from services import (
    registrar_auditoria,
    usuario_bloqueado, registrar_intento_fallido, resetear_intentos,
    generar_api_token, verificar_api_token, revocar_api_token,
    email_configurado, enviar_alertas_stock_bajo,
    s3_configurado, subir_imagen_bytes,
    generar_ordenes_compra_sugeridas,
    asegurar_codigo_producto, generar_qr_base64,
    generar_orden_desde_cotizacion,
    dias_atraso_solicitud, enviar_notificacion_solicitud, dias_gracia_historial_solicitud,
    sesion_activa_minutos, formatear_duracion,
)

load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", os.urandom(32))
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=8)

MAX_UPLOAD_MB = int(os.environ.get("MAX_UPLOAD_MB", 10))
app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_MB * 1024 * 1024

DEBUG = os.environ.get("DEBUG", "False").lower() == "true"
FORCE_HTTPS = os.environ.get("FORCE_HTTPS", "False").lower() == "true"

app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=FORCE_HTTPS,
)

if FORCE_HTTPS:
    from werkzeug.middleware.proxy_fix import ProxyFix
    app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

csrf = CSRFProtect(app)

ACTIVO_TRUE = "TRUE" if USE_POSTGRES else "1"
ACTIVO_FALSE = "FALSE" if USE_POSTGRES else "0"

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "static", "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "webp"}


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def save_uploaded_image(archivo):
    """Guarda imagen con compresion WebP. Usa S3/R2 si esta configurado, si no disco local."""
    if not archivo or not archivo.filename:
        return None
    if not allowed_file(archivo.filename):
        flash("Tipo de archivo no permitido (png, jpg, jpeg, gif, webp).", "warning")
        return None

    prefijo = datetime.now().strftime("%Y%m%d%H%M%S")
    try:
        from PIL import Image
        img = Image.open(archivo.stream)
        if img.mode in ("RGBA", "P"):
            img = img.convert("RGBA")
            background = Image.new("RGB", img.size, (255, 255, 255))
            background.paste(img, mask=img.split()[3])
            img = background
        elif img.mode != "RGB":
            img = img.convert("RGB")
        max_size = 900
        if max(img.width, img.height) > max_size:
            img.thumbnail((max_size, max_size), Image.LANCZOS)
        filename = f"{prefijo}_{secure_filename(archivo.filename.rsplit('.', 1)[0])}.webp"
        buf = io.BytesIO()
        img.save(buf, "WEBP", quality=82, method=6)
        buf.seek(0)
        return subir_imagen_bytes(buf.read(), filename, "image/webp")
    except Exception as e:
        app.logger.warning("Pillow no disponible, guardando sin comprimir: %s", e)
        filename = f"{prefijo}_{secure_filename(archivo.filename)}"
        archivo.seek(0)
        return subir_imagen_bytes(archivo.read(), filename, archivo.mimetype or "application/octet-stream")


ALLOWED_DOC_EXTENSIONS = {"pdf", "png", "jpg", "jpeg", "webp"}


def allowed_doc_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_DOC_EXTENSIONS


def guardar_documento_cotizacion(archivo):
    """Guarda el PDF/imagen de una cotizacion (S3/R2 si esta configurado, si no disco local)."""
    if not archivo or not archivo.filename:
        return None
    if not allowed_doc_file(archivo.filename):
        flash("Tipo de archivo no permitido para el documento (pdf, png, jpg, jpeg, webp).", "warning")
        return None
    prefijo = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"{prefijo}_{secure_filename(archivo.filename)}"
    content_type = archivo.mimetype or "application/octet-stream"
    archivo.seek(0)
    return subir_imagen_bytes(archivo.read(), filename, content_type, carpeta="documentos")


def safe_int(v, default=0):
    try:
        return max(0, min(int(v), 9_999_999))
    except (TypeError, ValueError):
        return default


def safe_float(v, default=0.0):
    try:
        return max(0.0, float(v))
    except (TypeError, ValueError):
        return default


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login", next=request.url))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        if session.get("role") != "admin":
            flash("No tienes permiso para realizar esta accion.", "danger")
            return redirect(url_for("index"))
        return f(*args, **kwargs)
    return decorated


def solicitante_o_admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        if session.get("role") not in ("admin", "solicitante"):
            flash("No tienes permiso para realizar esta acción.", "danger")
            return redirect(url_for("index"))
        return f(*args, **kwargs)
    return decorated


def comprador_o_admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        if session.get("role") not in ("admin", "comprador"):
            flash("No tienes permiso para realizar esta acción.", "danger")
            return redirect(url_for("listar_solicitudes"))
        return f(*args, **kwargs)
    return decorated


def api_auth(roles=None):
    """Autentica via sesion de navegador o token Bearer (Authorization: Bearer <token>)."""
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            auth_header = request.headers.get("Authorization", "")
            token = auth_header[7:].strip() if auth_header.lower().startswith("bearer ") else None
            actor = None
            if token:
                token_row = verificar_api_token(token)
                if not token_row:
                    return jsonify({"error": "Token invalido, inactivo o expirado."}), 401
                conn = get_db_connection()
                cur = conn.cursor()
                ph = p()
                cur.execute(f"SELECT id, role, nombre, username FROM usuarios WHERE id = {ph}", (token_row["usuario_id"],))
                u = cur.fetchone()
                conn.close()
                actor = {
                    "role": u["role"] if u else "viewer",
                    "nombre": (u["nombre"] or u["username"]) if u else "API",
                    "id": u["id"] if u else None,
                    "via": "token",
                }
            elif session.get("logged_in"):
                actor = {
                    "role": session.get("role"),
                    "nombre": session.get("nombre"),
                    "id": session.get("user_id"),
                    "via": "session",
                }
            else:
                return jsonify({"error": "No autenticado. Usa sesion iniciada o Authorization: Bearer <token>."}), 401

            if roles and actor["role"] not in roles:
                return jsonify({"error": "Permiso insuficiente para esta accion."}), 403

            request.api_actor = actor
            return f(*args, **kwargs)
        return decorated
    return decorator


@app.before_request
def setup_db():
    init_db()


@app.before_request
def forzar_https():
    if FORCE_HTTPS and request.headers.get("X-Forwarded-Proto", "http") == "http":
        url = request.url.replace("http://", "https://", 1)
        return redirect(url, code=301)


@app.before_request
def registrar_actividad_sesion():
    """Actualiza 'ultima_actividad' de la sesion actual (con throttle de 60s para no saturar la BD)."""
    if not session.get("logged_in") or not session.get("sesion_id"):
        return
    if request.endpoint in ("static",):
        return
    ahora_dt = datetime.now()
    ultima_guardada = session.get("_hb_ts")
    if ultima_guardada:
        try:
            if (ahora_dt - datetime.fromisoformat(ultima_guardada)).total_seconds() < 60:
                return
        except Exception:
            pass
    try:
        conn = get_db_connection()
        ph = p()
        conn.cursor().execute(
            f"UPDATE sesiones SET ultima_actividad = {ph} WHERE id = {ph} AND fin IS NULL",
            (ahora_dt.strftime("%Y-%m-%d %H:%M:%S"), session.get("sesion_id")),
        )
        conn.commit()
        conn.close()
        session["_hb_ts"] = ahora_dt.isoformat()
    except Exception:
        pass


@app.errorhandler(413)
def archivo_muy_grande(e):
    flash(f"El archivo supera el limite de {MAX_UPLOAD_MB} MB.", "danger")
    return redirect(request.referrer or url_for("listar_productos")), 413


@app.errorhandler(404)
def not_found(e):
    return render_template("404.html"), 404


# AUTH
@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("logged_in"):
        if session.get("role") == "comprador":
            return redirect(url_for("listar_solicitudes"))
        return redirect(url_for("index"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        conn = get_db_connection()
        cur = conn.cursor()
        ph = p()
        cur.execute(f"SELECT * FROM usuarios WHERE username = {ph} AND activo = {ACTIVO_TRUE}", (username,))
        usuario = cur.fetchone()
        conn.close()
        if usuario:
            bloqueado, minutos = usuario_bloqueado(usuario)
            if bloqueado:
                flash(f"Cuenta bloqueada temporalmente por multiples intentos fallidos. "
                      f"Intenta de nuevo en {minutos} minuto(s).", "danger")
                return render_template("login.html", next=request.args.get("next", ""))

            stored = usuario["password"]
            if isinstance(stored, str):
                stored_bytes = stored.encode()
            else:
                stored_bytes = stored

            password_ok = False

            if stored_bytes.startswith(b"$2b$") or stored_bytes.startswith(b"$2a$"):
                try:
                    password_ok = bcrypt.checkpw(password.encode(), stored_bytes)
                except Exception:
                    password_ok = False

            elif b":" in stored_bytes or b"$" in stored_bytes[:10]:
                try:
                    from werkzeug.security import check_password_hash
                    password_ok = check_password_hash(stored, password)
                except Exception:
                    password_ok = False
                if password_ok:
                    try:
                        new_hash = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
                        conn2 = get_db_connection()
                        ph2 = p()
                        conn2.cursor().execute(
                            f"UPDATE usuarios SET password = {ph2} WHERE id = {ph2}",
                            (new_hash, usuario["id"])
                        )
                        conn2.commit()
                        conn2.close()
                    except Exception:
                        pass

            else:
                import hmac as _hmac
                password_ok = _hmac.compare_digest(password, stored)
                if password_ok:
                    try:
                        new_hash = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
                        conn2 = get_db_connection()
                        ph2 = p()
                        conn2.cursor().execute(
                            f"UPDATE usuarios SET password = {ph2} WHERE id = {ph2}",
                            (new_hash, usuario["id"])
                        )
                        conn2.commit()
                        conn2.close()
                    except Exception:
                        pass

            if password_ok:
                resetear_intentos(usuario["id"])
                session.permanent = True
                session["logged_in"] = True
                session["username"] = username
                session["nombre"] = usuario["nombre"] or username
                session["role"] = usuario["role"]
                session["user_id"] = usuario["id"]

                ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ip_cliente = request.headers.get("X-Forwarded-For", request.remote_addr or "").split(",")[0].strip()
                user_agent = (request.headers.get("User-Agent", "") or "")[:255]

                conn3 = get_db_connection()
                cur3 = conn3.cursor()
                ph3 = p()
                cur3.execute(
                    f"UPDATE usuarios SET ultimo_login = {ph3}, ultima_ip = {ph3} WHERE id = {ph3}",
                    (ahora, ip_cliente, usuario["id"]),
                )
                cur3.execute(
                    f"""INSERT INTO sesiones (usuario_id, ip, user_agent, inicio, ultima_actividad)
                        VALUES ({ph3},{ph3},{ph3},{ph3},{ph3})""",
                    (usuario["id"], ip_cliente, user_agent, ahora, ahora),
                )
                conn3.commit()
                cur3.execute("SELECT last_insert_rowid() AS id" if not USE_POSTGRES else "SELECT lastval() AS id")
                session["sesion_id"] = cur3.fetchone()["id"]
                conn3.close()

                registrar_auditoria("usuarios", usuario["id"], "login", usuario["id"],
                                     usuario["nombre"] or username, "Inicio de sesion exitoso")
                destino_default = url_for("listar_solicitudes") if usuario["role"] == "comprador" else url_for("index")
                next_url = request.form.get("next") or destino_default
                return redirect(next_url)
            else:
                intentos_previos = usuario["failed_attempts"] if "failed_attempts" in usuario.keys() else 0
                registrar_intento_fallido(usuario["id"], intentos_previos)

        flash("Usuario o contrasena incorrectos.", "danger")
    return render_template("login.html", next=request.args.get("next", ""))


@app.route("/logout")
def logout():
    sesion_id = session.get("sesion_id")
    if sesion_id:
        try:
            conn = get_db_connection()
            cur = conn.cursor()
            ph = p()
            cur.execute(f"SELECT inicio FROM sesiones WHERE id = {ph}", (sesion_id,))
            fila = cur.fetchone()
            ahora_dt = datetime.now()
            ahora = ahora_dt.strftime("%Y-%m-%d %H:%M:%S")
            duracion = None
            if fila and fila["inicio"]:
                try:
                    inicio_dt = datetime.fromisoformat(str(fila["inicio"])[:19])
                    duracion = int((ahora_dt - inicio_dt).total_seconds())
                except Exception:
                    duracion = None
            cur.execute(
                f"UPDATE sesiones SET fin = {ph}, ultima_actividad = {ph}, duracion_segundos = {ph} WHERE id = {ph}",
                (ahora, ahora, duracion, sesion_id),
            )
            conn.commit()
            conn.close()
        except Exception:
            pass
    session.clear()
    flash("Sesion cerrada correctamente.", "info")
    return redirect(url_for("login"))


# DASHBOARD
@app.route("/")
@login_required
def index():
    if session.get("role") == "comprador":
        return redirect(url_for("listar_solicitudes"))
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(f"SELECT COUNT(*) AS n FROM productos WHERE activo = {ACTIVO_TRUE}")
    total_productos = cur.fetchone()["n"]
    cur.execute(f"SELECT COALESCE(SUM(stock_actual), 0) AS n FROM productos WHERE activo = {ACTIVO_TRUE}")
    total_stock = cur.fetchone()["n"]
    cur.execute(f"SELECT COUNT(*) AS n FROM productos WHERE stock_actual < stock_minimo AND activo = {ACTIVO_TRUE}")
    stock_bajo = cur.fetchone()["n"]
    cur.execute("""
        SELECT p.nombre, SUM(m.cantidad) AS total_salidas
        FROM movimientos m
        JOIN productos p ON p.id = m.producto_id
        WHERE m.tipo = 'salida'
        GROUP BY p.id, p.nombre
        ORDER BY total_salidas DESC
        LIMIT 5
    """)
    top_salidas = cur.fetchall()
    cur.execute("""
        SELECT date(fecha) AS dia,
               SUM(CASE WHEN tipo='entrada' THEN cantidad ELSE 0 END) AS entradas,
               SUM(CASE WHEN tipo='salida' THEN cantidad ELSE 0 END) AS salidas
        FROM movimientos
        WHERE date(fecha) >= date('now', '-13 days')
        GROUP BY dia ORDER BY dia
    """)
    filas_dias = cur.fetchall()
    cur.execute(f"""
        SELECT id, codigo, nombre, stock_actual, stock_minimo, categoria, equipo
        FROM productos WHERE stock_actual < stock_minimo AND activo = {ACTIVO_TRUE}
        ORDER BY (stock_minimo - stock_actual) DESC LIMIT 8
    """)
    alertas = cur.fetchall()
    conn.close()

    return render_template(
        "index.html",
        total_productos=total_productos,
        total_stock=total_stock,
        stock_bajo=stock_bajo,
        top_labels=[r["nombre"] for r in top_salidas],
        top_values=[r["total_salidas"] for r in top_salidas],
        dias=[f["dia"] for f in filas_dias],
        entradas=[f["entradas"] for f in filas_dias],
        salidas=[f["salidas"] for f in filas_dias],
        alertas=alertas,
        email_configurado=email_configurado(),
    )


# PRODUCTOS
@app.route("/productos")
@login_required
def listar_productos():
    if session.get("role") == "comprador":
        return redirect(url_for("listar_solicitudes"))
    q = request.args.get("q", "").strip()
    categoria_filtro = request.args.get("categoria", "").strip()
    proveedor_filtro = request.args.get("proveedor", "").strip()
    equipo_filtro = request.args.get("equipo", "").strip()
    stock_bajo_param = request.args.get("stock_bajo", "").strip()
    mostrar_inactivos = request.args.get("inactivos", "").strip() == "1" and session.get("role") == "admin"

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT DISTINCT categoria FROM productos WHERE categoria IS NOT NULL AND categoria <> '' ORDER BY categoria")
    categorias = [r["categoria"] for r in cur.fetchall()]
    cur.execute("SELECT DISTINCT proveedor FROM productos WHERE proveedor IS NOT NULL AND proveedor <> '' ORDER BY proveedor")
    proveedores = [r["proveedor"] for r in cur.fetchall()]
    cur.execute("SELECT DISTINCT equipo FROM productos WHERE equipo IS NOT NULL AND equipo <> '' ORDER BY equipo")
    equipos = [r["equipo"] for r in cur.fetchall()]

    sql = "SELECT * FROM productos WHERE 1=1"
    params = []

    if not mostrar_inactivos:
        sql += f" AND activo = {ACTIVO_TRUE}"

    if q:
        sql += " AND (codigo LIKE ? OR nombre LIKE ? OR categoria LIKE ? OR ubicacion LIKE ? OR proveedor LIKE ?)"
        pat = f"%{q}%"
        params.extend([pat] * 5)
    if categoria_filtro:
        sql += " AND categoria = ?"
        params.append(categoria_filtro)
    if proveedor_filtro:
        sql += " AND proveedor = ?"
        params.append(proveedor_filtro)
    if equipo_filtro:
        sql += " AND equipo = ?"
        params.append(equipo_filtro)
    if stock_bajo_param:
        sql += " AND stock_actual < stock_minimo"
    sql += " ORDER BY nombre"

    if p() == "%s":
        sql = sql.replace("?", "%s")

    cur.execute(sql, params)
    productos = cur.fetchall()
    conn.close()

    return render_template(
        "productos.html",
        productos=productos, q=q,
        categorias=categorias, proveedores=proveedores, equipos=equipos,
        categoria_filtro=categoria_filtro, proveedor_filtro=proveedor_filtro,
        equipo_filtro=equipo_filtro, stock_bajo=stock_bajo_param,
        mostrar_inactivos=mostrar_inactivos,
    )


@app.route("/productos/nuevo", methods=["GET", "POST"])
@login_required
@admin_required
def nuevo_producto():
    if request.method == "POST":
        nombre = (request.form.get("nombre") or "").strip()
        if not nombre:
            flash("El nombre del producto es obligatorio.", "danger")
            return render_template("producto_form.html", producto=None)

        codigo = (request.form.get("codigo") or "").strip() or None
        categoria = (request.form.get("categoria") or "").strip()
        equipo = (request.form.get("equipo") or "").strip()
        linea = (request.form.get("linea") or "").strip()
        descripcion = (request.form.get("descripcion") or "").strip()
        stock_minimo = safe_int(request.form.get("stock_minimo"))
        stock_actual = safe_int(request.form.get("stock_actual"))
        ubicacion = (request.form.get("ubicacion") or "").strip()
        proveedor = (request.form.get("proveedor") or "").strip()
        precio = safe_float(request.form.get("precio"))

        imagen_url = (request.form.get("imagen_url") or "").strip() or None
        url_subida = save_uploaded_image(request.files.get("imagen_archivo"))
        if url_subida:
            imagen_url = url_subida

        ph = p()
        try:
            conn = get_db_connection()
            cur = conn.cursor()
            cur.execute(
                f"""INSERT INTO productos (codigo,nombre,descripcion,categoria,equipo,linea,
                    stock_minimo,stock_actual,ubicacion,proveedor,precio,imagen_url)
                    VALUES ({ph},{ph},{ph},{ph},{ph},{ph},{ph},{ph},{ph},{ph},{ph},{ph})""",
                (codigo, nombre, descripcion, categoria, equipo, linea,
                 stock_minimo, stock_actual, ubicacion, proveedor, precio, imagen_url),
            )
            conn.commit()
            cur.execute(f"SELECT id FROM productos WHERE nombre = {ph} ORDER BY id DESC LIMIT 1", (nombre,))
            nuevo_id_row = cur.fetchone()
            nuevo_id = nuevo_id_row["id"] if nuevo_id_row else None
            registrar_auditoria("productos", nuevo_id, "crear", session.get("user_id"), session.get("nombre"),
                                 f"Producto '{nombre}' (codigo {codigo or '-'}) creado con stock {stock_actual}")
            flash("Producto creado correctamente.", "success")
            return redirect(url_for("listar_productos"))
        except Exception as e:
            app.logger.error("Error al crear producto: %s", e)
            flash("No se pudo crear el producto (¿codigo duplicado?).", "danger")
        finally:
            conn.close()

    return render_template("producto_form.html", producto=None)


@app.route("/productos/<int:pid>/editar", methods=["GET", "POST"])
@login_required
@admin_required
def editar_producto(pid):
    conn = get_db_connection()
    cur = conn.cursor()
    ph = p()
    cur.execute(f"SELECT * FROM productos WHERE id = {ph}", (pid,))
    producto = cur.fetchone()
    if not producto:
        conn.close()
        flash("Producto no encontrado.", "warning")
        return redirect(url_for("listar_productos"))

    if request.method == "POST":
        nombre = (request.form.get("nombre") or "").strip()
        if not nombre:
            conn.close()
            flash("El nombre es obligatorio.", "danger")
            return redirect(url_for("editar_producto", pid=pid))

        codigo = (request.form.get("codigo") or "").strip() or None
        categoria = (request.form.get("categoria") or "").strip()
        equipo = (request.form.get("equipo") or "").strip()
        linea = (request.form.get("linea") or "").strip()
        descripcion = (request.form.get("descripcion") or "").strip()
        stock_minimo = safe_int(request.form.get("stock_minimo"))
        stock_actual = safe_int(request.form.get("stock_actual"))
        ubicacion = (request.form.get("ubicacion") or "").strip()
        proveedor = (request.form.get("proveedor") or "").strip()
        precio = safe_float(request.form.get("precio"))

        imagen_url = (request.form.get("imagen_url") or "").strip() or producto["imagen_url"]
        url_subida = save_uploaded_image(request.files.get("imagen_archivo"))
        if url_subida:
            imagen_url = url_subida

        cur.execute(f"SELECT id FROM productos WHERE codigo = {ph} AND id != {ph}", (codigo, pid))
        if cur.fetchone():
            conn.close()
            flash("Ya existe otro producto con ese codigo.", "danger")
            return redirect(url_for("editar_producto", pid=pid))

        try:
            cambios = []
            if producto["stock_minimo"] != stock_minimo:
                cambios.append(f"stock_minimo {producto['stock_minimo']}->{stock_minimo}")
            if producto["stock_actual"] != stock_actual:
                cambios.append(f"stock_actual {producto['stock_actual']}->{stock_actual}")
            if float(producto["precio"] or 0) != precio:
                cambios.append(f"precio {producto['precio']}->{precio}")

            cur.execute(
                f"""UPDATE productos SET codigo={ph}, nombre={ph}, descripcion={ph},
                    categoria={ph}, equipo={ph}, linea={ph}, stock_minimo={ph},
                    stock_actual={ph}, ubicacion={ph}, proveedor={ph}, precio={ph},
                    imagen_url={ph} WHERE id={ph}""",
                (codigo, nombre, descripcion, categoria, equipo, linea,
                 stock_minimo, stock_actual, ubicacion, proveedor, precio, imagen_url, pid),
            )
            conn.commit()
            registrar_auditoria("productos", pid, "editar", session.get("user_id"), session.get("nombre"),
                                 "; ".join(cambios) if cambios else "Edicion sin cambios de stock/precio")
            flash("Producto actualizado.", "success")
        except Exception as e:
            app.logger.error("Error al actualizar: %s", e)
            flash("No se pudo actualizar el producto.", "danger")
        finally:
            conn.close()
        return redirect(url_for("listar_productos"))

    conn.close()
    return render_template("producto_form.html", producto=producto)


@app.route("/productos/<int:pid>")
@login_required
def detalle_producto(pid):
    if session.get("role") == "comprador":
        return redirect(url_for("listar_solicitudes"))
    conn = get_db_connection()
    cur = conn.cursor()
    ph = p()
    cur.execute(f"SELECT * FROM productos WHERE id = {ph}", (pid,))
    producto = cur.fetchone()
    if not producto:
        conn.close()
        flash("Producto no encontrado.", "warning")
        return redirect(url_for("listar_productos"))
    cur.execute(f"""
        SELECT * FROM movimientos WHERE producto_id = {ph}
        ORDER BY fecha DESC, id DESC LIMIT 50
    """, (pid,))
    movimientos = cur.fetchall()
    conn.close()
    return render_template("producto_detalle.html", producto=producto, movimientos=movimientos)


@app.route("/productos/<int:pid>/eliminar", methods=["POST"])
@login_required
@admin_required
def eliminar_producto(pid):
    conn = get_db_connection()
    ph = p()
    try:
        conn.cursor().execute(f"UPDATE productos SET activo = {ACTIVO_FALSE} WHERE id = {ph}", (pid,))
        conn.commit()
        registrar_auditoria("productos", pid, "eliminar", session.get("user_id"), session.get("nombre"),
                             "Producto desactivado (soft delete)")
        flash("Producto desactivado. Puedes restaurarlo desde 'Ver inactivos'.", "info")
    finally:
        conn.close()
    return redirect(url_for("listar_productos"))


@app.route("/productos/<int:pid>/restaurar", methods=["POST"])
@login_required
@admin_required
def restaurar_producto(pid):
    conn = get_db_connection()
    ph = p()
    try:
        conn.cursor().execute(f"UPDATE productos SET activo = {ACTIVO_TRUE} WHERE id = {ph}", (pid,))
        conn.commit()
        registrar_auditoria("productos", pid, "restaurar", session.get("user_id"), session.get("nombre"),
                             "Producto reactivado")
        flash("Producto restaurado.", "success")
    finally:
        conn.close()
    return redirect(url_for("listar_productos", inactivos=1))


# MOVIMIENTOS
@app.route("/movimientos")
@login_required
def listar_movimientos():
    if session.get("role") == "comprador":
        return redirect(url_for("listar_solicitudes"))
    tipo = request.args.get("tipo", "")
    producto_id = request.args.get("producto_id", "")
    desde = request.args.get("desde", "")
    hasta = request.args.get("hasta", "")

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, nombre, codigo FROM productos ORDER BY nombre")
    productos = cur.fetchall()

    sql = """
        SELECT m.*, p.nombre AS nombre_producto, p.codigo AS codigo_producto
        FROM movimientos m JOIN productos p ON p.id = m.producto_id WHERE 1=1
    """
    params = []
    ph = p()
    if tipo in ("entrada", "salida"):
        sql += f" AND m.tipo = {ph}"; params.append(tipo)
    if producto_id:
        sql += f" AND m.producto_id = {ph}"; params.append(producto_id)
    if desde:
        sql += f" AND date(m.fecha) >= date({ph})"; params.append(desde)
    if hasta:
        sql += f" AND date(m.fecha) <= date({ph})"; params.append(hasta)
    sql += " ORDER BY m.fecha DESC, m.id DESC"

    cur.execute(sql, params)
    movimientos = cur.fetchall()
    conn.close()

    return render_template(
        "movimientos.html",
        movimientos=movimientos, productos=productos,
        tipo=tipo, producto_id=producto_id, desde=desde, hasta=hasta,
    )


@app.route("/movimientos/nuevo/<tipo>", methods=["GET", "POST"])
@login_required
def nuevo_movimiento(tipo):
    if session.get("role") == "comprador":
        return redirect(url_for("listar_solicitudes"))
    if tipo not in ("entrada", "salida"):
        flash("Tipo de movimiento no valido.", "danger")
        return redirect(url_for("listar_movimientos"))

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(f"SELECT id, codigo, nombre, stock_actual, stock_minimo FROM productos WHERE activo = {ACTIVO_TRUE} ORDER BY nombre")
    productos = cur.fetchall()
    ph = p()

    if request.method == "POST":
        try:
            producto_id = int(request.form["producto_id"])
            cantidad = int(request.form["cantidad"])
        except (KeyError, ValueError):
            conn.close()
            flash("Datos invalidos.", "danger")
            return redirect(url_for("nuevo_movimiento", tipo=tipo))

        if cantidad <= 0:
            conn.close()
            flash("La cantidad debe ser mayor a cero.", "danger")
            return redirect(url_for("nuevo_movimiento", tipo=tipo))

        usuario = request.form.get("usuario", session.get("nombre", ""))[:100]
        motivo = request.form.get("motivo", "")[:500]

        cur.execute(f"SELECT stock_actual FROM productos WHERE id = {ph}", (producto_id,))
        prod = cur.fetchone()
        if not prod:
            conn.close()
            flash("Producto no encontrado.", "danger")
            return redirect(url_for("nuevo_movimiento", tipo=tipo))

        stock_actual = prod["stock_actual"]
        if tipo == "salida" and cantidad > stock_actual:
            conn.close()
            flash(f"Stock insuficiente. Solo hay {stock_actual} unidades.", "danger")
            return redirect(url_for("nuevo_movimiento", tipo=tipo))

        nuevo_stock = stock_actual + cantidad if tipo == "entrada" else stock_actual - cantidad
        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        user_id = session.get("user_id")

        cur.execute(
            f"INSERT INTO movimientos (producto_id,tipo,cantidad,fecha,usuario,usuario_id,motivo) VALUES ({ph},{ph},{ph},{ph},{ph},{ph},{ph})",
            (producto_id, tipo, cantidad, fecha, usuario, user_id, motivo),
        )
        cur.execute(f"UPDATE productos SET stock_actual = {ph} WHERE id = {ph}", (nuevo_stock, producto_id))
        conn.commit()
        registrar_auditoria("movimientos", producto_id, tipo, user_id, usuario,
                             f"{cantidad} unidad(es). Stock {stock_actual}->{nuevo_stock}. Motivo: {motivo or '-'}")
        conn.close()
        flash("Movimiento registrado.", "success")
        return redirect(url_for("listar_movimientos"))

    conn.close()
    return render_template("movimiento_form.html", tipo=tipo, productos=productos)


# EXPORTAR EXCEL
def _estilo_encabezado(ws):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    thin = Side(style="thin")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_h = PatternFill("solid", fgColor="1B4F8A")
    font_h = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill = fill_h; cell.font = font_h; cell.alignment = center; cell.border = brd
    return brd


@app.route("/exportar/stock_bajo")
@login_required
@admin_required
def exportar_stock_bajo():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(f"""
        SELECT codigo, nombre, proveedor, ubicacion,
               (stock_minimo - stock_actual) AS cantidad_a_comprar
        FROM productos WHERE stock_actual < stock_minimo AND activo = {ACTIVO_TRUE}
        AND (stock_minimo - stock_actual) > 0 ORDER BY proveedor, nombre
    """)
    filas = cur.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Faltantes"
    headers = ["Codigo", "Producto", "Proveedor", "Ubicacion", "Cant. a Comprar"]
    ws.append(headers)
    brd = _estilo_encabezado(ws)
    for fila in filas:
        ws.append(list(fila))
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = brd
        row[-1].font = Font(bold=True); row[-1].alignment = Alignment(horizontal="center")
    for i, w in enumerate([15, 38, 22, 18, 18], 1):
        ws.column_dimensions[chr(64+i)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, download_name=f"stock_bajo_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/exportar/pedidos")
@login_required
@admin_required
def exportar_pedidos():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(f"""
        SELECT proveedor, codigo, nombre, ubicacion,
               (stock_minimo - stock_actual) AS cantidad
        FROM productos WHERE stock_actual < stock_minimo AND activo = {ACTIVO_TRUE}
        AND (stock_minimo - stock_actual) > 0 ORDER BY proveedor, nombre
    """)
    filas = cur.fetchall()
    conn.close()

    pedidos = defaultdict(list)
    for f in filas:
        pedidos[f["proveedor"] or "Sin proveedor"].append(f)

    thin = Side(style="thin")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    zip_buf = io.BytesIO()

    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zipf:
        for proveedor, items in pedidos.items():
            wb = Workbook(); ws = wb.active; ws.title = "Pedido"
            ws.append(["Codigo", "Producto", "Ubicacion", "Cantidad"])
            for c in ws[1]:
                c.fill = PatternFill("solid", fgColor="1B4F8A")
                c.font = Font(color="FFFFFF", bold=True)
                c.alignment = Alignment(horizontal="center")
                c.border = brd
            for it in items:
                ws.append([it["codigo"], it["nombre"], it["ubicacion"], it["cantidad"]])
            for row in ws.iter_rows(min_row=2):
                for c in row: c.border = brd
                row[-1].alignment = Alignment(horizontal="center"); row[-1].font = Font(bold=True)
            for col, w in zip(["A","B","C","D"], [15, 42, 20, 15]):
                ws.column_dimensions[col].width = w
            buf = io.BytesIO(); wb.save(buf); buf.seek(0)
            nombre_prov = "".join(c for c in proveedor if c.isalnum() or c in " _-")[:30]
            zipf.writestr(f"Pedido_{nombre_prov}.xlsx", buf.read())

    zip_buf.seek(0)
    return send_file(zip_buf, download_name=f"Pedidos_{datetime.now().strftime('%Y-%m-%d')}.zip",
                     as_attachment=True, mimetype="application/zip")


@app.route("/exportar/valorizacion")
@login_required
@admin_required
def exportar_valorizacion():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(f"""
        SELECT codigo, nombre, categoria, proveedor, stock_actual, precio,
               (stock_actual * precio) AS valor
        FROM productos WHERE activo = {ACTIVO_TRUE} ORDER BY valor DESC
    """)
    filas = cur.fetchall()
    conn.close()

    total = sum((f["valor"] or 0) for f in filas)

    wb = Workbook(); ws = wb.active; ws.title = "Valorizacion"
    ws.append(["Codigo", "Producto", "Categoria", "Proveedor", "Stock", "Precio Unit.", "Valor Total"])
    brd = _estilo_encabezado(ws)
    for f in filas:
        ws.append([f["codigo"], f["nombre"], f["categoria"], f["proveedor"],
                   f["stock_actual"], f["precio"], f["valor"]])
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.border = brd
        row[5].number_format = '#,##0'
        row[6].number_format = '#,##0'
    fila_total = ws.max_row + 1
    ws.cell(row=fila_total, column=6, value="TOTAL:").font = Font(bold=True)
    ws.cell(row=fila_total, column=6).alignment = Alignment(horizontal="right")
    celda_total = ws.cell(row=fila_total, column=7, value=total)
    celda_total.font = Font(bold=True)
    celda_total.number_format = '#,##0'
    for i, w in enumerate([15, 38, 18, 22, 10, 14, 16], 1):
        ws.column_dimensions[chr(64+i)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, download_name=f"valorizacion_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/exportar/rotacion")
@login_required
@admin_required
def exportar_rotacion():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    dias = safe_int(request.args.get("dias", 90), 90) or 90

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(f"SELECT id, codigo, nombre, stock_actual FROM productos WHERE activo = {ACTIVO_TRUE}")
    productos = cur.fetchall()
    cur.execute("SELECT producto_id, tipo, cantidad, fecha FROM movimientos")
    movs = cur.fetchall()
    conn.close()

    cutoff = datetime.now() - timedelta(days=dias)
    salidas_por_producto = defaultdict(int)
    for m in movs:
        if m["tipo"] != "salida":
            continue
        fecha_raw = m["fecha"]
        try:
            fecha_dt = fecha_raw if isinstance(fecha_raw, datetime) else datetime.fromisoformat(str(fecha_raw)[:19])
        except Exception:
            continue
        if fecha_dt >= cutoff:
            salidas_por_producto[m["producto_id"]] += (m["cantidad"] or 0)

    filas = []
    for prod in productos:
        salidas = salidas_por_producto.get(prod["id"], 0)
        rotacion = round(salidas / prod["stock_actual"], 2) if prod["stock_actual"] else 0
        filas.append((prod["codigo"], prod["nombre"], prod["stock_actual"], salidas, rotacion))
    filas.sort(key=lambda r: r[3], reverse=True)

    wb = Workbook(); ws = wb.active; ws.title = "Rotacion"
    ws.append(["Codigo", "Producto", "Stock actual", f"Salidas ({dias} dias)", "Indice de rotacion"])
    brd = _estilo_encabezado(ws)
    for f in filas:
        ws.append(list(f))
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.border = brd
        row[2].alignment = Alignment(horizontal="center")
        row[3].alignment = Alignment(horizontal="center")
        row[4].alignment = Alignment(horizontal="center")
    for i, w in enumerate([15, 38, 14, 20, 20], 1):
        ws.column_dimensions[chr(64+i)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, download_name=f"rotacion_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/exportar/abc")
@login_required
@admin_required
def exportar_abc():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(f"""
        SELECT codigo, nombre, stock_actual, precio, (stock_actual * precio) AS valor
        FROM productos WHERE activo = {ACTIVO_TRUE} ORDER BY valor DESC
    """)
    filas = cur.fetchall()
    conn.close()

    total = sum((f["valor"] or 0) for f in filas) or 1
    acumulado = 0
    clasificadas = []
    for f in filas:
        acumulado += (f["valor"] or 0)
        pct_acumulado = acumulado / total * 100
        clase = "A" if pct_acumulado <= 80 else ("B" if pct_acumulado <= 95 else "C")
        clasificadas.append((f["codigo"], f["nombre"], f["stock_actual"], f["valor"], round(pct_acumulado, 1), clase))

    colores = {"A": "C6EFCE", "B": "FFEB9C", "C": "FFC7CE"}

    wb = Workbook(); ws = wb.active; ws.title = "Clasificacion ABC"
    ws.append(["Codigo", "Producto", "Stock", "Valor", "% Acumulado", "Clase"])
    brd = _estilo_encabezado(ws)
    for f in clasificadas:
        ws.append(list(f))
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.border = brd
        clase_val = row[5].value
        row[5].fill = PatternFill("solid", fgColor=colores.get(clase_val, "FFFFFF"))
        row[5].alignment = Alignment(horizontal="center")
        row[5].font = Font(bold=True)
        row[2].alignment = Alignment(horizontal="center")
        row[4].alignment = Alignment(horizontal="center")
    for i, w in enumerate([15, 38, 10, 16, 14, 10], 1):
        ws.column_dimensions[chr(64+i)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, download_name=f"clasificacion_abc_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ETIQUETAS QR
@app.route("/productos/<int:pid>/etiqueta")
@login_required
@admin_required
def etiqueta_producto(pid):
    conn = get_db_connection()
    cur = conn.cursor()
    ph = p()
    cur.execute(f"SELECT * FROM productos WHERE id = {ph}", (pid,))
    producto = cur.fetchone()
    conn.close()
    if not producto:
        flash("Producto no encontrado.", "warning")
        return redirect(url_for("listar_productos"))

    codigo = asegurar_codigo_producto(pid, producto["codigo"])
    qr_data_uri = generar_qr_base64(codigo)
    return render_template("etiqueta.html", producto=producto, codigo=codigo, qr_data_uri=qr_data_uri)


@app.route("/productos/etiquetas")
@login_required
@admin_required
def etiquetas_lote():
    ids_param = request.args.get("ids", "").strip()
    solo_stock_bajo = request.args.get("stock_bajo", "").strip() == "1"

    conn = get_db_connection()
    cur = conn.cursor()
    ph = p()

    if ids_param:
        ids = [int(i) for i in ids_param.split(",") if i.strip().isdigit()]
        if not ids:
            productos = []
        else:
            placeholders = ",".join([ph] * len(ids))
            cur.execute(f"SELECT * FROM productos WHERE id IN ({placeholders}) ORDER BY nombre", ids)
            productos = cur.fetchall()
    elif solo_stock_bajo:
        cur.execute(f"SELECT * FROM productos WHERE stock_actual < stock_minimo AND activo = {ACTIVO_TRUE} ORDER BY nombre")
        productos = cur.fetchall()
    else:
        cur.execute(f"SELECT * FROM productos WHERE activo = {ACTIVO_TRUE} ORDER BY nombre")
        productos = cur.fetchall()
    conn.close()

    etiquetas = []
    for prod in productos:
        codigo = asegurar_codigo_producto(prod["id"], prod["codigo"])
        etiquetas.append({
            "id": prod["id"], "nombre": prod["nombre"], "codigo": codigo,
            "ubicacion": prod["ubicacion"], "qr_data_uri": generar_qr_base64(codigo),
        })

    return render_template("etiquetas_lote.html", etiquetas=etiquetas)


# ORDENES DE COMPRA
@app.route("/ordenes_compra")
@login_required
@admin_required
def listar_ordenes_compra():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM ordenes_compra ORDER BY id DESC")
    ordenes = cur.fetchall()
    conn.close()
    return render_template("ordenes_compra.html", ordenes=ordenes)


@app.route("/ordenes_compra/generar", methods=["POST"])
@login_required
@admin_required
def generar_ordenes_compra():
    ids = generar_ordenes_compra_sugeridas(session.get("user_id"), session.get("nombre"))
    if ids:
        registrar_auditoria("ordenes_compra", None, "generar", session.get("user_id"), session.get("nombre"),
                             f"{len(ids)} orden(es) generadas automaticamente")
        flash(f"Se generaron {len(ids)} orden(es) de compra sugeridas.", "success")
    else:
        flash("No hay productos bajo el stock minimo en este momento.", "info")
    return redirect(url_for("listar_ordenes_compra"))


@app.route("/ordenes_compra/<int:oid>/estado", methods=["POST"])
@login_required
@admin_required
def cambiar_estado_orden(oid):
    nuevo_estado = request.form.get("estado", "pendiente")
    if nuevo_estado not in ("pendiente", "enviada", "recibida", "cancelada"):
        nuevo_estado = "pendiente"
    conn = get_db_connection()
    ph = p()
    conn.cursor().execute(f"UPDATE ordenes_compra SET estado = {ph} WHERE id = {ph}", (nuevo_estado, oid))
    conn.commit()
    conn.close()
    registrar_auditoria("ordenes_compra", oid, "cambiar_estado", session.get("user_id"), session.get("nombre"),
                         f"Nuevo estado: {nuevo_estado}")
    flash("Estado de la orden actualizado.", "success")
    return redirect(url_for("listar_ordenes_compra"))


@app.route("/ordenes_compra/<int:oid>/exportar")
@login_required
@admin_required
def exportar_orden_compra(oid):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    conn = get_db_connection()
    cur = conn.cursor()
    ph = p()
    cur.execute(f"SELECT * FROM ordenes_compra WHERE id = {ph}", (oid,))
    orden = cur.fetchone()
    if not orden:
        conn.close()
        flash("Orden de compra no encontrada.", "warning")
        return redirect(url_for("listar_ordenes_compra"))
    cur.execute(f"SELECT * FROM ordenes_compra_items WHERE orden_id = {ph}", (oid,))
    items = cur.fetchall()
    conn.close()

    wb = Workbook(); ws = wb.active; ws.title = "Orden de Compra"
    ws.merge_cells("A1:E1")
    ws["A1"] = "ORDEN DE COMPRA -- INVENTARIO WINTEC"
    ws["A1"].font = Font(bold=True, size=14, color="1B4F8A")
    ws["A3"] = "N Orden:"; ws["B3"] = orden["numero"]
    ws["A4"] = "Fecha:"; ws["B4"] = str(orden["fecha"])[:19]
    ws["A5"] = "Proveedor:"; ws["B5"] = orden["proveedor_nombre"]
    ws["A6"] = "Estado:"; ws["B6"] = orden["estado"]
    ws["A7"] = "Generado por:"; ws["B7"] = orden["creado_por"] or "-"
    for r in range(3, 8):
        ws.cell(row=r, column=1).font = Font(bold=True)

    header_row = 9
    ws.cell(row=header_row, column=1, value="Codigo")
    ws.cell(row=header_row, column=2, value="Producto")
    ws.cell(row=header_row, column=3, value="Cantidad")
    ws.cell(row=header_row, column=4, value="Precio Unit.")
    ws.cell(row=header_row, column=5, value="Subtotal")
    for c in ws[header_row]:
        c.fill = PatternFill("solid", fgColor="1B4F8A")
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center")

    row_i = header_row + 1
    for it in items:
        ws.cell(row=row_i, column=1, value=it["codigo"])
        ws.cell(row=row_i, column=2, value=it["nombre"])
        ws.cell(row=row_i, column=3, value=it["cantidad"]).alignment = Alignment(horizontal="center")
        ws.cell(row=row_i, column=4, value=it["precio_unitario"])
        ws.cell(row=row_i, column=5, value=it["subtotal"])
        row_i += 1

    ws.cell(row=row_i + 1, column=4, value="TOTAL ESTIMADO:").font = Font(bold=True)
    ws.cell(row=row_i + 1, column=5, value=orden["total_estimado"]).font = Font(bold=True)

    for i, w in enumerate([15, 42, 12, 14, 14], 1):
        ws.column_dimensions[chr(64+i)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, download_name=f"{orden['numero']}.xlsx",
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# UBICACIONES (bodegas / cajas / estantes)
@app.route("/ubicaciones")
@login_required
def listar_ubicaciones():
    if session.get("role") == "comprador":
        return redirect(url_for("listar_solicitudes"))
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM ubicaciones ORDER BY nombre")
    ubicaciones = cur.fetchall()

    resultado = []
    for u in ubicaciones:
        ph = p()
        cur.execute(
            f"SELECT COUNT(*) AS n FROM productos WHERE ubicacion = {ph} AND activo = {ACTIVO_TRUE}",
            (u["nombre"],),
        )
        n = cur.fetchone()["n"]
        resultado.append({"id": u["id"], "nombre": u["nombre"], "n_productos": n})
    conn.close()
    return render_template("ubicaciones.html", ubicaciones=resultado)


@app.route("/ubicaciones/<int:uid>")
@login_required
def ver_ubicacion(uid):
    if session.get("role") == "comprador":
        return redirect(url_for("listar_solicitudes"))
    conn = get_db_connection()
    cur = conn.cursor()
    ph = p()
    cur.execute(f"SELECT * FROM ubicaciones WHERE id = {ph}", (uid,))
    ubicacion = cur.fetchone()
    if not ubicacion:
        conn.close()
        flash("Ubicación no encontrada.", "warning")
        return redirect(url_for("listar_ubicaciones"))

    cur.execute(
        f"""SELECT * FROM productos WHERE ubicacion = {ph} AND activo = {ACTIVO_TRUE} ORDER BY nombre""",
        (ubicacion["nombre"],),
    )
    productos = cur.fetchall()
    conn.close()
    return render_template("ubicacion_detalle.html", ubicacion=ubicacion, productos=productos)


@app.route("/ubicaciones/<int:uid>/etiqueta")
@login_required
@admin_required
def etiqueta_ubicacion(uid):
    conn = get_db_connection()
    cur = conn.cursor()
    ph = p()
    cur.execute(f"SELECT * FROM ubicaciones WHERE id = {ph}", (uid,))
    ubicacion = cur.fetchone()
    conn.close()
    if not ubicacion:
        flash("Ubicación no encontrada.", "warning")
        return redirect(url_for("listar_ubicaciones"))

    url_ubicacion = url_for("ver_ubicacion", uid=uid, _external=True)
    qr_data_uri = generar_qr_base64(url_ubicacion, box_size=10)
    return render_template("etiqueta_ubicacion.html", ubicacion=ubicacion, qr_data_uri=qr_data_uri, url_ubicacion=url_ubicacion)


@app.route("/ubicaciones/etiquetas")
@login_required
@admin_required
def etiquetas_ubicaciones_lote():
    ids_param = request.args.get("ids", "").strip()
    conn = get_db_connection()
    cur = conn.cursor()
    ph = p()

    if ids_param:
        ids = [int(i) for i in ids_param.split(",") if i.strip().isdigit()]
        if ids:
            placeholders = ",".join([ph] * len(ids))
            cur.execute(f"SELECT * FROM ubicaciones WHERE id IN ({placeholders}) ORDER BY nombre", ids)
        else:
            cur.execute("SELECT * FROM ubicaciones WHERE 0")
    else:
        cur.execute("SELECT * FROM ubicaciones ORDER BY nombre")
    ubicaciones = cur.fetchall()
    conn.close()

    etiquetas = []
    for u in ubicaciones:
        url_ubicacion = url_for("ver_ubicacion", uid=u["id"], _external=True)
        etiquetas.append({
            "nombre": u["nombre"],
            "qr_data_uri": generar_qr_base64(url_ubicacion, box_size=10),
        })
    return render_template("etiquetas_ubicaciones_lote.html", etiquetas=etiquetas)


# COTIZACIONES
@app.route("/cotizaciones")
@login_required
@admin_required
def listar_cotizaciones():
    proveedor_filtro = request.args.get("proveedor", "").strip()
    estado_filtro = request.args.get("estado", "").strip()

    conn = get_db_connection()
    cur = conn.cursor()
    ph = p()
    sql = "SELECT * FROM cotizaciones WHERE 1=1"
    params = []
    if proveedor_filtro:
        sql += f" AND proveedor_nombre = {ph}"
        params.append(proveedor_filtro)
    if estado_filtro:
        sql += f" AND estado = {ph}"
        params.append(estado_filtro)
    sql += " ORDER BY fecha_recibida DESC, id DESC"
    cur.execute(sql, params)
    cotizaciones = cur.fetchall()

    cur.execute("SELECT nombre FROM proveedores ORDER BY nombre")
    proveedores = [r["nombre"] for r in cur.fetchall()]
    conn.close()

    return render_template(
        "cotizaciones.html", cotizaciones=cotizaciones, proveedores=proveedores,
        proveedor_filtro=proveedor_filtro, estado_filtro=estado_filtro,
    )


@app.route("/cotizaciones/nueva", methods=["GET", "POST"])
@login_required
@admin_required
def nueva_cotizacion():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT nombre FROM proveedores ORDER BY nombre")
    proveedores = [r["nombre"] for r in cur.fetchall()]

    if request.method == "POST":
        proveedor_nombre = (request.form.get("proveedor") or "").strip()
        fecha_recibida = request.form.get("fecha_recibida") or datetime.now().strftime("%Y-%m-%d")
        fecha_vigencia = request.form.get("fecha_vigencia") or None
        notas = (request.form.get("notas") or "").strip()

        nombres_items = request.form.getlist("nombre_item[]")
        codigos_items = request.form.getlist("codigo_item[]")
        cantidades_items = request.form.getlist("cantidad_item[]")
        precios_items = request.form.getlist("precio_item[]")

        items = []
        total_calculado = 0.0
        for i in range(len(nombres_items)):
            nombre_it = (nombres_items[i] or "").strip()
            if not nombre_it:
                continue
            cantidad = safe_int(cantidades_items[i] if i < len(cantidades_items) else 0, 0)
            precio = safe_float(precios_items[i] if i < len(precios_items) else 0, 0.0)
            subtotal = cantidad * precio
            total_calculado += subtotal
            items.append({
                "nombre": nombre_it,
                "codigo": (codigos_items[i] if i < len(codigos_items) else "").strip(),
                "cantidad": cantidad, "precio_unitario": precio, "subtotal": subtotal,
            })

        monto_total_form = request.form.get("monto_total", "").strip()
        monto_total = safe_float(monto_total_form) if monto_total_form else total_calculado

        documento_url = guardar_documento_cotizacion(request.files.get("documento"))

        ph = p()
        numero = _siguiente_numero_cotizacion_helper(cur)
        cur.execute(
            f"""INSERT INTO cotizaciones (numero, proveedor_nombre, fecha_recibida, fecha_vigencia,
                monto_total, estado, documento_url, notas, creado_por, creado_por_id)
                VALUES ({ph},{ph},{ph},{ph},{ph},'pendiente',{ph},{ph},{ph},{ph})""",
            (numero, proveedor_nombre, fecha_recibida, fecha_vigencia,
             monto_total, documento_url, notas, session.get("nombre"), session.get("user_id")),
        )
        conn.commit()
        cur.execute(f"SELECT id FROM cotizaciones WHERE numero = {ph}", (numero,))
        cot_id = cur.fetchone()["id"]

        for it in items:
            cur.execute(
                f"""INSERT INTO cotizacion_items (cotizacion_id, codigo, nombre, cantidad, precio_unitario, subtotal)
                    VALUES ({ph},{ph},{ph},{ph},{ph},{ph})""",
                (cot_id, it["codigo"], it["nombre"], it["cantidad"], it["precio_unitario"], it["subtotal"]),
            )
        conn.commit()
        conn.close()

        registrar_auditoria("cotizaciones", cot_id, "crear", session.get("user_id"), session.get("nombre"),
                             f"Cotización {numero} de '{proveedor_nombre}' por {monto_total}")
        flash(f"Cotización {numero} registrada.", "success")
        return redirect(url_for("detalle_cotizacion", cid=cot_id))

    conn.close()
    return render_template("cotizacion_form.html", proveedores=proveedores,
                            hoy=datetime.now().strftime("%Y-%m-%d"))


@app.route("/cotizaciones/<int:cid>")
@login_required
@admin_required
def detalle_cotizacion(cid):
    conn = get_db_connection()
    cur = conn.cursor()
    ph = p()
    cur.execute(f"SELECT * FROM cotizaciones WHERE id = {ph}", (cid,))
    cotizacion = cur.fetchone()
    if not cotizacion:
        conn.close()
        flash("Cotización no encontrada.", "warning")
        return redirect(url_for("listar_cotizaciones"))
    cur.execute(f"SELECT * FROM cotizacion_items WHERE cotizacion_id = {ph}", (cid,))
    items = cur.fetchall()
    conn.close()
    return render_template("cotizacion_detalle.html", cotizacion=cotizacion, items=items)


@app.route("/cotizaciones/<int:cid>/estado", methods=["POST"])
@login_required
@admin_required
def cambiar_estado_cotizacion(cid):
    nuevo_estado = request.form.get("estado", "pendiente")
    if nuevo_estado not in ("pendiente", "aceptada", "rechazada", "vencida", "cancelada"):
        nuevo_estado = "pendiente"

    conn = get_db_connection()
    cur = conn.cursor()
    ph = p()
    cur.execute(f"UPDATE cotizaciones SET estado = {ph} WHERE id = {ph}", (nuevo_estado, cid))
    conn.commit()
    conn.close()

    registrar_auditoria("cotizaciones", cid, "cambiar_estado", session.get("user_id"), session.get("nombre"),
                         f"Nuevo estado: {nuevo_estado}")

    orden_id = None
    if nuevo_estado == "aceptada":
        orden_id = generar_orden_desde_cotizacion(cid, session.get("user_id"), session.get("nombre"))

    if orden_id:
        registrar_auditoria("ordenes_compra", orden_id, "generar", session.get("user_id"), session.get("nombre"),
                             f"Generada automáticamente desde cotización #{cid}")
        flash("Cotización aceptada. Se generó la orden de compra correspondiente.", "success")
    else:
        flash("Estado de la cotización actualizado.", "success")
    return redirect(url_for("detalle_cotizacion", cid=cid))


def _siguiente_numero_cotizacion_helper(cur):
    from services import _siguiente_numero_cotizacion
    return _siguiente_numero_cotizacion(cur)


# SOLICITUDES DEL EQUIPO
@app.route("/solicitudes")
@login_required
def listar_solicitudes():
    # Visible para todos los roles logueados (admin, solicitante, viewer).
    estado_filtro = request.args.get("estado", "").strip()
    vista = request.args.get("vista", "activo").strip()
    if vista not in ("activo", "historial"):
        vista = "activo"

    conn = get_db_connection()
    cur = conn.cursor()
    ph = p()
    sql = "SELECT * FROM solicitudes WHERE 1=1"
    params = []

    if session.get("role") == "solicitante":
        sql += f" AND solicitado_por_id = {ph}"
        params.append(session.get("user_id"))

    dias_gracia = dias_gracia_historial_solicitud()
    if vista == "historial":
        sql += " AND estado IN ('comprado','rechazado','cancelado')"
    else:
        limite = (datetime.now() - timedelta(days=dias_gracia)).strftime("%Y-%m-%d %H:%M:%S")
        sql += " AND estado != 'comprado'"
        sql += f" AND (estado NOT IN ('rechazado','cancelado') OR fecha_atendida IS NULL OR fecha_atendida >= {ph})"
        params.append(limite)

    if estado_filtro:
        sql += f" AND estado = {ph}"
        params.append(estado_filtro)

    orden_campo = "fecha_atendida" if vista == "historial" else "fecha_solicitud"
    sql += f" ORDER BY {orden_campo} DESC, id DESC"

    cur.execute(sql, params)
    filas = cur.fetchall()
    conn.close()

    umbral = dias_atraso_solicitud()
    ahora = datetime.now()
    solicitudes = []
    for s in filas:
        s_dict = dict(s)
        atrasada = False
        if s_dict["estado"] == "pendiente" and s_dict["fecha_solicitud"]:
            try:
                fecha_dt = datetime.fromisoformat(str(s_dict["fecha_solicitud"])[:19])
                if (ahora - fecha_dt).days >= umbral:
                    atrasada = True
            except Exception:
                pass
        s_dict["atrasada"] = atrasada
        solicitudes.append(s_dict)

    return render_template("solicitudes.html", solicitudes=solicitudes, estado_filtro=estado_filtro,
                            umbral_dias=umbral, vista=vista, dias_gracia=dias_gracia)


@app.route("/solicitudes/nueva", methods=["GET", "POST"])
@login_required
@solicitante_o_admin_required
def nueva_solicitud():
    if request.method == "POST":
        producto_id = request.form.get("producto_id") or None
        nombre_item = (request.form.get("nombre_item") or "").strip()
        if not nombre_item:
            flash("Debes indicar qué necesitas.", "danger")
            return redirect(url_for("nueva_solicitud"))

        descripcion = (request.form.get("descripcion") or "").strip()
        cantidad = safe_int(request.form.get("cantidad"), 1) or 1
        urgencia = request.form.get("urgencia", "normal")
        if urgencia not in ("normal", "urgente"):
            urgencia = "normal"

        foto_url = save_uploaded_image(request.files.get("foto"))

        ph = p()
        fecha_ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            f"""INSERT INTO solicitudes (producto_id, nombre_item, descripcion, cantidad, urgencia,
                foto_url, estado, solicitado_por, solicitado_por_id, fecha_solicitud)
                VALUES ({ph},{ph},{ph},{ph},{ph},{ph},'pendiente',{ph},{ph},{ph})""",
            (producto_id, nombre_item, descripcion, cantidad, urgencia,
             foto_url, session.get("nombre"), session.get("user_id"), fecha_ahora),
        )
        conn.commit()
        cur.execute("SELECT last_insert_rowid() AS id" if not USE_POSTGRES else "SELECT lastval() AS id")
        sol_id = cur.fetchone()["id"]
        conn.close()

        registrar_auditoria("solicitudes", sol_id, "crear", session.get("user_id"), session.get("nombre"),
                             f"Solicitó {cantidad}x '{nombre_item}' (urgencia: {urgencia})")

        ok, _msg = enviar_notificacion_solicitud({
            "solicitado_por": session.get("nombre"), "nombre_item": nombre_item,
            "cantidad": cantidad, "urgencia": urgencia, "descripcion": descripcion,
        })

        flash("Solicitud registrada." + (" Se avisó por correo al comprador." if ok else ""), "success")
        return redirect(url_for("listar_solicitudes"))

    return render_template("solicitud_form.html")


@app.route("/solicitudes/<int:sid>")
@login_required
def detalle_solicitud(sid):
    if session.get("role") not in ("admin", "solicitante", "viewer", "comprador"):
        flash("No tienes acceso a esta sección.", "warning")
        return redirect(url_for("index"))

    conn = get_db_connection()
    cur = conn.cursor()
    ph = p()
    cur.execute(f"SELECT * FROM solicitudes WHERE id = {ph}", (sid,))
    solicitud = cur.fetchone()
    conn.close()

    if not solicitud:
        flash("Solicitud no encontrada.", "warning")
        return redirect(url_for("listar_solicitudes"))
    if session.get("role") == "solicitante" and solicitud["solicitado_por_id"] != session.get("user_id"):
        flash("Esa solicitud no es tuya.", "warning")
        return redirect(url_for("listar_solicitudes"))

    return render_template("solicitud_detalle.html", solicitud=solicitud)


@app.route("/solicitudes/<int:sid>/estado", methods=["POST"])
@login_required
@comprador_o_admin_required
def cambiar_estado_solicitud(sid):
    nuevo_estado = request.form.get("estado", "pendiente")
    if nuevo_estado not in ("pendiente", "comprado", "rechazado", "cancelado"):
        nuevo_estado = "pendiente"

    ph = p()
    conn = get_db_connection()
    fecha_atendida = datetime.now().strftime("%Y-%m-%d %H:%M:%S") if nuevo_estado in ("comprado", "rechazado", "cancelado") else None
    conn.cursor().execute(
        f"UPDATE solicitudes SET estado = {ph}, fecha_atendida = {ph}, comprado_por = {ph} WHERE id = {ph}",
        (nuevo_estado, fecha_atendida, session.get("nombre") if nuevo_estado == "comprado" else None, sid),
    )
    conn.commit()
    conn.close()

    registrar_auditoria("solicitudes", sid, "cambiar_estado", session.get("user_id"), session.get("nombre"),
                         f"Nuevo estado: {nuevo_estado}")
    flash("Estado de la solicitud actualizado.", "success")
    return redirect(url_for("detalle_solicitud", sid=sid))


# ADMIN -- USUARIOS
@app.route("/admin/usuarios")
@admin_required
def admin_usuarios():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, username, nombre, role, activo, created_at FROM usuarios ORDER BY id")
    usuarios = cur.fetchall()
    conn.close()
    return render_template("admin_usuarios.html", usuarios=usuarios)


@app.route("/admin/usuarios/nuevo", methods=["POST"])
@admin_required
def crear_usuario():
    username = (request.form.get("username") or "").strip()
    password = request.form.get("password") or ""
    nombre = (request.form.get("nombre") or "").strip()
    role = request.form.get("role", "viewer")

    if not username or not password:
        flash("Usuario y contrasena son obligatorios.", "danger")
        return redirect(url_for("admin_usuarios"))

    if role not in ("admin", "viewer", "solicitante", "comprador"):
        role = "viewer"

    hashed = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
    ph = p()
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            f"INSERT INTO usuarios (username, password, nombre, role) VALUES ({ph},{ph},{ph},{ph})",
            (username, hashed, nombre, role)
        )
        conn.commit()
        cur.execute(f"SELECT id FROM usuarios WHERE username = {ph}", (username,))
        row = cur.fetchone()
        registrar_auditoria("usuarios", row["id"] if row else None, "crear",
                             session.get("user_id"), session.get("nombre"), f"Usuario '{username}' ({role}) creado")
        flash(f"Usuario '{username}' creado.", "success")
    except Exception as e:
        app.logger.error("Error crear usuario: %s", e)
        flash("Ya existe un usuario con ese nombre.", "danger")
    finally:
        conn.close()
    return redirect(url_for("admin_usuarios"))


@app.route("/admin/usuarios/<int:uid>/toggle", methods=["POST"])
@admin_required
def toggle_usuario(uid):
    if uid == session.get("user_id"):
        flash("No puedes desactivar tu propia cuenta.", "warning")
        return redirect(url_for("admin_usuarios"))
    conn = get_db_connection()
    ph = p()
    cur = conn.cursor()
    cur.execute(f"SELECT activo FROM usuarios WHERE id = {ph}", (uid,))
    u = cur.fetchone()
    if u:
        nuevo = 0 if u["activo"] else 1
        cur.execute(f"UPDATE usuarios SET activo = {ph} WHERE id = {ph}", (nuevo, uid))
        conn.commit()
        registrar_auditoria("usuarios", uid, "toggle", session.get("user_id"), session.get("nombre"),
                             f"Usuario {'activado' if nuevo else 'desactivado'}")
        flash("Estado de usuario actualizado.", "info")
    conn.close()
    return redirect(url_for("admin_usuarios"))


@app.route("/admin/usuarios/<int:uid>/reset_password", methods=["POST"])
@admin_required
def reset_password(uid):
    nueva = request.form.get("nueva_password", "")
    if len(nueva) < 6:
        flash("La contrasena debe tener al menos 6 caracteres.", "danger")
        return redirect(url_for("admin_usuarios"))
    hashed = bcrypt.hashpw(nueva.encode(), bcrypt.gensalt()).decode()
    ph = p()
    conn = get_db_connection()
    conn.cursor().execute(f"UPDATE usuarios SET password = {ph}, failed_attempts = 0, locked_until = NULL WHERE id = {ph}", (hashed, uid))
    conn.commit()
    conn.close()
    registrar_auditoria("usuarios", uid, "reset_password", session.get("user_id"), session.get("nombre"),
                         "Contrasena restablecida por admin")
    flash("Contrasena actualizada.", "success")
    return redirect(url_for("admin_usuarios"))


# ADMIN -- CATEGORIAS, EQUIPOS, PROVEEDORES
@app.route("/admin/categorias")
@admin_required
def admin_categorias():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM categorias ORDER BY nombre")
    items = cur.fetchall()
    conn.close()
    return render_template("admin_catalogos.html", items=items, tipo="categorias", titulo="Categorias")


@app.route("/admin/equipos")
@admin_required
def admin_equipos():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM equipos ORDER BY nombre")
    items = cur.fetchall()
    conn.close()
    return render_template("admin_catalogos.html", items=items, tipo="equipos", titulo="Equipos")


@app.route("/admin/ubicaciones")
@admin_required
def admin_ubicaciones():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM ubicaciones ORDER BY nombre")
    items = cur.fetchall()
    conn.close()
    return render_template("admin_catalogos.html", items=items, tipo="ubicaciones", titulo="Ubicaciones")


@app.route("/admin/proveedores")
@admin_required
def admin_proveedores():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM proveedores ORDER BY nombre")
    items = cur.fetchall()
    conn.close()
    return render_template("admin_proveedores.html", items=items)


@app.route("/admin/<tipo>/nuevo", methods=["POST"])
@admin_required
def admin_nuevo_catalogo(tipo):
    if tipo not in ("categorias", "equipos", "ubicaciones"):
        return redirect(url_for("index"))
    nombre = (request.form.get("nombre") or "").strip()
    if not nombre:
        flash("El nombre es obligatorio.", "danger")
        return redirect(url_for(f"admin_{tipo}"))
    ph = p()
    try:
        conn = get_db_connection()
        ignore = "OR IGNORE" if ph == "?" else ""
        conflict = "ON CONFLICT DO NOTHING" if ph == "%s" else ""
        conn.cursor().execute(f"INSERT {ignore} INTO {tipo} (nombre) VALUES ({ph}) {conflict}".strip(), (nombre,))
        conn.commit()
        registrar_auditoria(tipo, None, "crear", session.get("user_id"), session.get("nombre"), f"'{nombre}' agregado")
        flash(f"'{nombre}' agregado.", "success")
    except Exception as e:
        app.logger.error(e)
        flash("Error al agregar.", "danger")
    finally:
        conn.close()
    return redirect(url_for(f"admin_{tipo}"))


@app.route("/admin/<tipo>/<int:item_id>/eliminar", methods=["POST"])
@admin_required
def admin_eliminar_catalogo(tipo, item_id):
    if tipo not in ("categorias", "equipos", "ubicaciones"):
        return redirect(url_for("index"))
    ph = p()
    conn = get_db_connection()
    try:
        conn.cursor().execute(f"DELETE FROM {tipo} WHERE id = {ph}", (item_id,))
        conn.commit()
        registrar_auditoria(tipo, item_id, "eliminar", session.get("user_id"), session.get("nombre"))
        flash("Elemento eliminado.", "info")
    finally:
        conn.close()
    return redirect(url_for(f"admin_{tipo}"))


@app.route("/admin/proveedores/nuevo", methods=["POST"])
@admin_required
def nuevo_proveedor():
    nombre = (request.form.get("nombre") or "").strip()
    if not nombre:
        flash("Nombre obligatorio.", "danger")
        return redirect(url_for("admin_proveedores"))
    ph = p()
    try:
        conn = get_db_connection()
        ignore = "OR IGNORE" if ph == "?" else ""
        conflict = "ON CONFLICT DO NOTHING" if ph == "%s" else ""
        conn.cursor().execute(
            f"INSERT {ignore} INTO proveedores (nombre, contacto, email, telefono) VALUES ({ph},{ph},{ph},{ph}) {conflict}".strip(),
            (nombre, request.form.get("contacto",""), request.form.get("email",""), request.form.get("telefono",""))
        )
        conn.commit()
        registrar_auditoria("proveedores", None, "crear", session.get("user_id"), session.get("nombre"), f"Proveedor '{nombre}' agregado")
        flash(f"Proveedor '{nombre}' agregado.", "success")
    except Exception as e:
        app.logger.error(e)
        flash("Error al agregar proveedor.", "danger")
    finally:
        conn.close()
    return redirect(url_for("admin_proveedores"))


@app.route("/admin/proveedores/<int:pid>/eliminar", methods=["POST"])
@admin_required
def eliminar_proveedor(pid):
    ph = p()
    conn = get_db_connection()
    try:
        conn.cursor().execute(f"DELETE FROM proveedores WHERE id = {ph}", (pid,))
        conn.commit()
        registrar_auditoria("proveedores", pid, "eliminar", session.get("user_id"), session.get("nombre"))
        flash("Proveedor eliminado.", "info")
    finally:
        conn.close()
    return redirect(url_for("admin_proveedores"))


# ADMIN -- AUDITORIA
@app.route("/admin/auditoria")
@admin_required
def admin_auditoria():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM auditoria ORDER BY id DESC LIMIT 300")
    logs = cur.fetchall()
    conn.close()
    return render_template("admin_auditoria.html", logs=logs)


# ADMIN -- CONTROL DE ACCESOS
@app.route("/admin/accesos")
@admin_required
def admin_accesos():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT id, username, nombre, role, activo, ultimo_login, ultima_ip FROM usuarios ORDER BY username")
    usuarios = [dict(u) for u in cur.fetchall()]

    cur.execute("""
        SELECT s.usuario_id, s.ip, s.ultima_actividad, s.fin
        FROM sesiones s
        INNER JOIN (
            SELECT usuario_id, MAX(inicio) AS max_inicio FROM sesiones GROUP BY usuario_id
        ) ult ON s.usuario_id = ult.usuario_id AND s.inicio = ult.max_inicio
    """)
    ultimas_sesiones = {row["usuario_id"]: dict(row) for row in cur.fetchall()}

    umbral_minutos = sesion_activa_minutos()
    ahora = datetime.now()
    for u in usuarios:
        ult = ultimas_sesiones.get(u["id"])
        en_linea = False
        if ult and not ult["fin"] and ult["ultima_actividad"]:
            try:
                actividad_dt = datetime.fromisoformat(str(ult["ultima_actividad"])[:19])
                if (ahora - actividad_dt).total_seconds() <= umbral_minutos * 60:
                    en_linea = True
            except Exception:
                pass
        u["en_linea"] = en_linea
        u["ip_reciente"] = ult["ip"] if ult else None

    cur.execute("""
        SELECT s.*, u.username, u.nombre
        FROM sesiones s
        LEFT JOIN usuarios u ON s.usuario_id = u.id
        ORDER BY s.inicio DESC
        LIMIT 100
    """)
    historial = []
    for row in cur.fetchall():
        h = dict(row)
        if h.get("duracion_segundos") is not None:
            h["duracion_texto"] = formatear_duracion(h["duracion_segundos"])
        elif h.get("ultima_actividad") and h.get("inicio"):
            try:
                inicio_dt = datetime.fromisoformat(str(h["inicio"])[:19])
                act_dt = datetime.fromisoformat(str(h["ultima_actividad"])[:19])
                h["duracion_texto"] = formatear_duracion(int((act_dt - inicio_dt).total_seconds())) + " (en curso)"
            except Exception:
                h["duracion_texto"] = "—"
        else:
            h["duracion_texto"] = "—"
        historial.append(h)

    conn.close()
    return render_template("admin_accesos.html", usuarios=usuarios, historial=historial, umbral_minutos=umbral_minutos)


# ADMIN -- TOKENS DE API
@app.route("/admin/api_tokens")
@admin_required
def admin_api_tokens():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT t.id, t.nombre, t.activo, t.created_at, t.last_used_at, u.username
        FROM api_tokens t LEFT JOIN usuarios u ON u.id = t.usuario_id
        ORDER BY t.id DESC
    """)
    tokens = cur.fetchall()
    conn.close()
    return render_template("admin_api_tokens.html", tokens=tokens)


@app.route("/admin/api_tokens/nuevo", methods=["POST"])
@admin_required
def crear_api_token_route():
    nombre = (request.form.get("nombre") or "Token API").strip() or "Token API"
    token = generar_api_token(session.get("user_id"), nombre)
    registrar_auditoria("api_tokens", None, "crear", session.get("user_id"), session.get("nombre"), f"Token '{nombre}' creado")
    flash(f"Token creado -- copialo ahora, no se volvera a mostrar: {token}", "success")
    return redirect(url_for("admin_api_tokens"))


@app.route("/admin/api_tokens/<int:tid>/revocar", methods=["POST"])
@admin_required
def revocar_token_route(tid):
    revocar_api_token(tid)
    registrar_auditoria("api_tokens", tid, "revocar", session.get("user_id"), session.get("nombre"))
    flash("Token revocado.", "info")
    return redirect(url_for("admin_api_tokens"))


# ADMIN -- SISTEMA (backups + alertas email)
@app.route("/admin/sistema")
@admin_required
def admin_sistema():
    backups_dir = os.path.join(BASE_DIR, "backups")
    backups = []
    if os.path.isdir(backups_dir):
        backups = sorted(os.listdir(backups_dir), reverse=True)[:20]
    return render_template(
        "admin_sistema.html",
        backups=backups,
        email_ok=email_configurado(),
        s3_ok=s3_configurado(),
        force_https=FORCE_HTTPS,
        use_postgres=USE_POSTGRES,
    )


@app.route("/admin/backup", methods=["POST"])
@admin_required
def ejecutar_backup():
    import backup_db
    try:
        ruta = backup_db.backup_postgres() if USE_POSTGRES else backup_db.backup_sqlite()
        backup_db.limpiar_backups_antiguos()
        if ruta:
            registrar_auditoria("sistema", None, "backup", session.get("user_id"), session.get("nombre"),
                                 os.path.basename(ruta))
            flash(f"Backup creado: {os.path.basename(ruta)}", "success")
        else:
            flash("No se pudo crear el backup. Revisa la consola del servidor.", "warning")
    except Exception as e:
        flash(f"Error al respaldar: {e}", "danger")
    return redirect(url_for("admin_sistema"))


@app.route("/admin/backup/<path:filename>/descargar")
@admin_required
def descargar_backup(filename):
    filename = secure_filename(filename)
    backups_dir = os.path.join(BASE_DIR, "backups")
    filepath = os.path.join(backups_dir, filename)
    if not os.path.isfile(filepath):
        flash("Archivo de backup no encontrado.", "warning")
        return redirect(url_for("admin_sistema"))
    return send_file(filepath, as_attachment=True)


@app.route("/admin/alertas/enviar", methods=["POST"])
@admin_required
def enviar_alerta_manual():
    ok, msg = enviar_alertas_stock_bajo()
    registrar_auditoria("alertas", None, "enviar_email", session.get("user_id"), session.get("nombre"), msg)
    flash(msg, "success" if ok else "warning")
    return redirect(url_for("admin_sistema"))


# API REST (autenticada con sesion o token Bearer)
@csrf.exempt
@app.route("/api/productos", methods=["GET"])
@api_auth()
def api_productos():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(f"SELECT * FROM productos WHERE activo = {ACTIVO_TRUE} ORDER BY nombre")
    productos = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify(productos)


@csrf.exempt
@app.route("/api/productos/<int:pid>", methods=["GET"])
@api_auth()
def api_detalle_producto(pid):
    conn = get_db_connection()
    cur = conn.cursor()
    ph = p()
    cur.execute(f"SELECT * FROM productos WHERE id = {ph}", (pid,))
    row = cur.fetchone()
    conn.close()
    if not row:
        return jsonify({"error": "Producto no encontrado"}), 404
    return jsonify(dict(row))


@csrf.exempt
@app.route("/api/productos", methods=["POST"])
@api_auth(roles=("admin",))
def api_crear_producto():
    data = request.get_json(silent=True) or request.form
    nombre = (data.get("nombre") or "").strip()
    if not nombre:
        return jsonify({"error": "nombre es obligatorio"}), 400
    codigo = (data.get("codigo") or "").strip() or None
    stock_minimo = safe_int(data.get("stock_minimo"))
    stock_actual = safe_int(data.get("stock_actual"))
    precio = safe_float(data.get("precio"))
    ph = p()
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            f"""INSERT INTO productos (codigo, nombre, categoria, equipo, proveedor,
                stock_minimo, stock_actual, ubicacion, precio)
                VALUES ({ph},{ph},{ph},{ph},{ph},{ph},{ph},{ph},{ph})""",
            (codigo, nombre, data.get("categoria", ""), data.get("equipo", ""),
             data.get("proveedor", ""), stock_minimo, stock_actual, data.get("ubicacion", ""), precio),
        )
        conn.commit()
        cur.execute(f"SELECT id FROM productos WHERE nombre = {ph} ORDER BY id DESC LIMIT 1", (nombre,))
        row = cur.fetchone()
        conn.close()
        registrar_auditoria("productos", row["id"] if row else None, "crear",
                             request.api_actor["id"], request.api_actor["nombre"], f"Creado via API: '{nombre}'")
        return jsonify({"ok": True, "id": row["id"] if row else None}), 201
    except Exception as e:
        return jsonify({"error": f"No se pudo crear ({e})"}), 400


@csrf.exempt
@app.route("/api/movimientos", methods=["POST"])
@api_auth(roles=("admin",))
def api_crear_movimiento():
    data = request.get_json(silent=True) or request.form
    tipo = data.get("tipo")
    if tipo not in ("entrada", "salida"):
        return jsonify({"error": "tipo debe ser 'entrada' o 'salida'"}), 400
    try:
        producto_id = int(data.get("producto_id"))
        cantidad = int(data.get("cantidad"))
    except (TypeError, ValueError):
        return jsonify({"error": "producto_id y cantidad son obligatorios y numericos"}), 400
    if cantidad <= 0:
        return jsonify({"error": "cantidad debe ser mayor a cero"}), 400

    conn = get_db_connection()
    cur = conn.cursor()
    ph = p()
    cur.execute(f"SELECT stock_actual FROM productos WHERE id = {ph}", (producto_id,))
    prod = cur.fetchone()
    if not prod:
        conn.close()
        return jsonify({"error": "Producto no encontrado"}), 404
    stock_actual = prod["stock_actual"]
    if tipo == "salida" and cantidad > stock_actual:
        conn.close()
        return jsonify({"error": f"Stock insuficiente. Solo hay {stock_actual} unidades."}), 400

    nuevo_stock = stock_actual + cantidad if tipo == "entrada" else stock_actual - cantidad
    fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    motivo = str(data.get("motivo", ""))[:500]
    usuario = request.api_actor["nombre"]

    cur.execute(
        f"INSERT INTO movimientos (producto_id,tipo,cantidad,fecha,usuario,usuario_id,motivo) VALUES ({ph},{ph},{ph},{ph},{ph},{ph},{ph})",
        (producto_id, tipo, cantidad, fecha, usuario, request.api_actor["id"], motivo),
    )
    cur.execute(f"UPDATE productos SET stock_actual = {ph} WHERE id = {ph}", (nuevo_stock, producto_id))
    conn.commit()
    conn.close()
    registrar_auditoria("movimientos", producto_id, tipo, request.api_actor["id"], usuario,
                         f"{cantidad} unidad(es) via API. Stock {stock_actual}->{nuevo_stock}")
    return jsonify({"ok": True, "stock_actual": nuevo_stock}), 201


# PWA -- manifest y service worker
@app.route("/manifest.json")
def manifest():
    return jsonify({
        "name": "Inventario Wintec",
        "short_name": "Wintec Inv.",
        "start_url": "/",
        "scope": "/",
        "display": "standalone",
        "background_color": "#0d1117",
        "theme_color": "#1B4F8A",
        "icons": [
            {"src": "/static/logo_wintec.png", "sizes": "192x192", "type": "image/png"},
            {"src": "/static/logo_wintec.png", "sizes": "512x512", "type": "image/png"},
        ],
    })


@app.route("/sw.js")
def service_worker():
    js = """
const CACHE_NAME = 'wintec-inv-v1';
const OFFLINE_URLS = ['/static/styles.css', '/static/logo_wintec.png'];

self.addEventListener('install', (event) => {
  event.waitUntil(caches.open(CACHE_NAME).then((cache) => cache.addAll(OFFLINE_URLS)));
  self.skipWaiting();
});

self.addEventListener('activate', (event) => {
  event.waitUntil(self.clients.claim());
});

self.addEventListener('fetch', (event) => {
  if (event.request.method !== 'GET') return;
  event.respondWith(
    fetch(event.request).catch(() => caches.match(event.request))
  );
});
"""
    return Response(js, mimetype="application/javascript")


@app.context_processor
def inject_globals():
    import os as _os
    logo_path = _os.path.join(BASE_DIR, "static", "logo_wintec.png")
    logo_exists = _os.path.exists(logo_path)
    return dict(logo_path=logo_exists, logo_exists=logo_exists)


if __name__ == "__main__":
    app.run(host="0.0.0.0", debug=DEBUG, port=5002)
